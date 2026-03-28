"""
CSV and Excel file parsers.
Both return the same ParseResult shape so callers don't care about format.
"""

import csv
import io
from collections import Counter
from dataclasses import dataclass, field

import openpyxl


@dataclass
class SheetInfo:
    name: str
    suggested_order: int  # lower = process first
    likely_entity_type: str | None  # asset | location | employee | vendor | None


@dataclass
class ParseResult:
    headers: list[str]
    preview_rows: list[dict]       # first 10 rows, {header: value}
    total_rows: int | None         # None if not determinable without full parse
    sheets: list[SheetInfo] = field(default_factory=list)  # Excel only

@dataclass
class ParseResultAll:
    headers: list[str]
    rows: list[dict]               # all rows, {header: value}
    total_rows: int | None         # None if not determinable without full parse
    sheets: list[SheetInfo] = field(default_factory=list)  # Excel only


# Keywords used to guess entity type from sheet name / column names
_ENTITY_KEYWORDS: dict[str, list[str]] = {
    "location": ["location", "lokacija", "building", "room", "floor", "rack", "warehouse"],
    "employee": ["employee", "zaposlenik", "person", "staff", "worker", "hr"],
    "vendor": ["vendor", "supplier", "dobavljac", "dobavljač", "company", "oib", "vat"],
    "asset": ["asset", "imovina", "serial", "inventory", "device", "laptop", "server"],
}


class CsvParser:
    """
    Parse CSV files with auto-delimiter detection.
    Priority: ; → , → \t
    """

    def _decode(self, content: bytes, encoding: str) -> str:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            return content.decode("latin-1")

    def _parse_rows(self, text: str) -> tuple[list[str], list[dict], int]:
        delimiter = self._detect_delimiter(text)
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
        return headers, rows, len(rows)

    def parse(self, content: bytes, encoding: str = "utf-8") -> ParseResult:
        text = self._decode(content, encoding)
        headers, rows, total = self._parse_rows(text)
        return ParseResult(headers=headers, preview_rows=rows[:10], total_rows=total)

    def parse_all(self, content: bytes, encoding: str = "utf-8") -> ParseResultAll:
        text = self._decode(content, encoding)
        headers, rows, total = self._parse_rows(text)
        return ParseResultAll(headers=headers, rows=rows, total_rows=total)

    def _detect_delimiter(self, text: str) -> str:
        """
        Try ; , \t in order. Score each by column consistency across first 20 rows.
        Higher consistency + higher column count wins. Priority order breaks ties.
        """
        lines = [l for l in text.split("\n")[:20] if l.strip()]
        if not lines:
            return ","

        best_delim = ","
        best_score = 0

        for delim in [";", ",", "\t"]:
            counts = [len(line.split(delim)) for line in lines]
            if not counts or max(counts) <= 1:
                continue
            # Score: rows sharing the modal column count × that count
            mode_count = Counter(counts).most_common(1)[0][0]
            consistency = sum(1 for c in counts if c == mode_count)
            score = consistency * mode_count
            if score > best_score:
                best_score = score
                best_delim = delim

        return best_delim


class ExcelParser:
    """
    Parse .xlsx files with openpyxl.
    Detects multiple sheets and suggests processing order.
    """

    def parse(self, content: bytes, sheet_name: str | None = None) -> ParseResult:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet_name]
            return self._parse_sheet(ws)

        # Analyse all sheets, return SheetInfo list + first-sheet preview
        sheets = self._analyse_sheets(wb)
        first_sheet_name = min(sheets, key=lambda s: s.suggested_order).name
        result = self._parse_sheet(wb[first_sheet_name])
        result.sheets = sheets
        return result

    def parse_all(self, content: bytes, sheet_name: str | None = None) -> ParseResultAll:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet_name]
            return self._parse_sheet_all(ws)

        # Analyse all sheets, return SheetInfo list + first-sheet preview
        sheets = self._analyse_sheets(wb)
        first_sheet_name = min(sheets, key=lambda s: s.suggested_order).name
        result = self._parse_sheet_all(wb[first_sheet_name])
        result.sheets = sheets
        return result

    def _parse_sheet(self, ws) -> ParseResult:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ParseResult(headers=[], preview_rows=[], total_rows=0)

        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        data_rows = rows[1:]
        total = len(data_rows)

        preview_rows = []
        for row in data_rows[:10]:
            preview_rows.append({
                headers[i]: (str(cell) if cell is not None else "")
                for i, cell in enumerate(row)
                if i < len(headers)
            })

        return ParseResult(headers=headers, preview_rows=preview_rows, total_rows=total)

    def _parse_sheet_all(self, ws) -> ParseResultAll:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ParseResultAll(headers=[], rows=[], total_rows=0)

        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        data_rows = rows[1:]
        total = len(data_rows)

        data_dicts = []
        for row in data_rows:
            data_dicts.append({
                headers[i]: (str(cell) if cell is not None else "")
                for i, cell in enumerate(row)
                if i < len(headers)
            })

        return ParseResultAll(headers=headers, rows=data_dicts, total_rows=total)

    def _analyse_sheets(self, wb) -> list[SheetInfo]:
        sheets = []
        all_names = wb.sheetnames

        for i, name in enumerate(all_names):
            ws = wb[name]
            # Read only first row for header analysis
            first_row = next(ws.iter_rows(values_only=True, max_row=1), ())
            headers = [str(c).lower() for c in first_row if c is not None]
            entity_type = self._detect_entity_type(name, headers)

            # Locations always first — everything else may reference them
            if entity_type == "location":
                suggested_order = 1
            elif entity_type == "employee":
                suggested_order = 2
            elif entity_type == "vendor":
                suggested_order = 3
            else:
                suggested_order = 10 + i  # assets and unknowns last

            sheets.append(SheetInfo(
                name=name,
                suggested_order=suggested_order,
                likely_entity_type=entity_type,
            ))

        sheets.sort(key=lambda s: s.suggested_order)
        return sheets

    def _detect_entity_type(self, sheet_name: str, headers: list[str]) -> str | None:
        combined = sheet_name.lower() + " " + " ".join(headers)
        scores: dict[str, int] = {}
        for entity, keywords in _ENTITY_KEYWORDS.items():
            scores[entity] = sum(1 for kw in keywords if kw in combined)
        best = max(scores, key=lambda k: scores[k])
        return best if scores[best] > 0 else None

    def _detect_dependencies(self, headers: list[str], all_sheet_names: list[str]) -> list[str]:
        """Return sheet names that this sheet likely references via a column name match."""
        deps = []
        for sheet in all_sheet_names:
            sheet_lower = sheet.lower()
            if any(sheet_lower in h.lower() for h in headers):
                deps.append(sheet)
        return deps
